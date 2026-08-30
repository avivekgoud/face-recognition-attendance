import io
import csv
from datetime import datetime, date
from typing import List, Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from ..models import AttendanceRecord
from ..config import settings

class ReportService:
    def __init__(self):
        pass

    def generate_csv(self, records: List[AttendanceRecord], org_name: str = "") -> bytes:
        """Generates a UTF-8 CSV export with BOM for Excel compatibility."""
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Header comments
        writer.writerow([f"Attendance Report - {org_name or settings.ORGANIZATION_NAME}"])
        writer.writerow([f"Generated On: {datetime.now().strftime('%Y-%m-%d %I:%M %p')}"])
        writer.writerow([]) # blank row
        
        # Table Columns
        writer.writerow([
            "Record ID",
            "Person ID / Reg No",
            "Full Name",
            "Department / Class",
            "Designation / Role",
            "Date",
            "Check-In Time",
            "Check-Out Time",
            "Status",
            "Confidence Score (%)",
            "Liveness Score",
            "Verification Mode",
            "Manual Override Reason",
            "Notes"
        ])

        for r in records:
            p = r.person
            dept = p.department.name if p and p.department else "N/A"
            writer.writerow([
                r.id,
                p.identifier if p else "N/A",
                p.full_name if p else "Unknown",
                dept,
                p.designation if p else "",
                r.date.strftime("%Y-%m-%d") if r.date else "",
                r.check_in_time.strftime("%I:%M:%S %p") if r.check_in_time else "--",
                r.check_out_time.strftime("%I:%M:%S %p") if r.check_out_time else "--",
                r.status.value,
                f"{r.recognition_confidence * 100:.1f}%",
                f"{r.liveness_score:.2f}",
                r.verification_mode.value,
                r.modification_reason or "",
                r.notes or ""
            ])

        # Return as UTF-8 bytes with BOM
        return output.getvalue().encode('utf-8-sig')

    def generate_excel(
        self,
        records: List[AttendanceRecord],
        org_name: str = "",
        filters_summary: Optional[Dict[str, Any]] = None
    ) -> bytes:
        """Generates a beautifully styled multi-sheet Excel spreadsheet with summary KPIs."""
        wb = Workbook()
        
        # Sheet 1: Detailed Logs
        ws = wb.active
        ws.title = "Attendance Logs"
        ws.views.sheetView[0].showGridLines = True

        # Palettes
        NAVY_FILL = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        WHITE_BOLD = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        TITLE_FONT = Font(name="Segoe UI", size=16, bold=True, color="0F172A")
        SUBTITLE_FONT = Font(name="Segoe UI", size=10, italic=True, color="64748B")
        REGULAR_FONT = Font(name="Segoe UI", size=10, color="1E293B")
        
        PRESENT_FILL = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        PRESENT_FONT = Font(name="Segoe UI", size=10, bold=True, color="166534")
        
        LATE_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
        LATE_FONT = Font(name="Segoe UI", size=10, bold=True, color="92400E")
        
        ABSENT_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        ABSENT_FONT = Font(name="Segoe UI", size=10, bold=True, color="991B1B")

        thin_side = Side(border_style="thin", color="CBD5E1")
        border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        # Header Title
        ws.merge_cells("A1:K1")
        ws["A1"] = f"Face Recognition Attendance Report — {org_name or settings.ORGANIZATION_NAME}"
        ws["A1"].font = TITLE_FONT
        
        ws.merge_cells("A2:K2")
        ws["A2"] = f"Exported on {datetime.now().strftime('%B %d, %Y at %I:%M %p')} | Total Records: {len(records)}"
        ws["A2"].font = SUBTITLE_FONT

        # Table Column Headers
        headers = [
            "ID", "Person ID", "Full Name", "Department / Class", "Designation",
            "Date", "Check-In", "Check-Out", "Status", "Confidence", "Mode"
        ]
        
        row_idx = 4
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=h)
            cell.fill = NAVY_FILL
            cell.font = WHITE_BOLD
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border_all
        ws.row_dimensions[row_idx].height = 25

        # Populate Data Rows
        for r in records:
            row_idx += 1
            p = r.person
            dept = p.department.name if p and p.department else "N/A"
            
            row_data = [
                r.id,
                p.identifier if p else "N/A",
                p.full_name if p else "Unknown",
                dept,
                p.designation if p else "",
                r.date.strftime("%Y-%m-%d") if r.date else "",
                r.check_in_time.strftime("%I:%M %p") if r.check_in_time else "--",
                r.check_out_time.strftime("%I:%M %p") if r.check_out_time else "--",
                r.status.value,
                f"{r.recognition_confidence * 100:.1f}%",
                r.verification_mode.value
            ]
            
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = REGULAR_FONT
                cell.border = border_all
                cell.alignment = Alignment(vertical="center", horizontal="center" if col_idx in [1, 2, 6, 7, 8, 9, 10, 11] else "left")
                
                # Highlight status column
                if col_idx == 9:
                    if val == "PRESENT":
                        cell.fill = PRESENT_FILL
                        cell.font = PRESENT_FONT
                    elif val == "LATE":
                        cell.fill = LATE_FILL
                        cell.font = LATE_FONT
                    elif val == "ABSENT":
                        cell.fill = ABSENT_FILL
                        cell.font = ABSENT_FONT
            
            ws.row_dimensions[row_idx].height = 20

        # Auto-adjust column widths
        from openpyxl.utils import get_column_letter
        for col_idx, col in enumerate(ws.columns, 1):
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        # Save to buffer
        out_buf = io.BytesIO()
        wb.save(out_buf)
        return out_buf.getvalue()

    def generate_pdf(
        self,
        records: List[AttendanceRecord],
        org_name: str = "",
        filters_summary: Optional[Dict[str, Any]] = None
    ) -> bytes:
        """Generates a professional PDF attendance report document using ReportLab."""
        buffer = io.BytesIO()
        # Landscape letter for optimal wide table viewing
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(letter),
            rightMargin=36,
            leftMargin=36,
            topMargin=36,
            bottomMargin=36
        )

        styles = getSampleStyleSheet()
        
        # Custom Paragraph styles
        title_style = ParagraphStyle(
            'DocTitle',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=18,
            leading=22,
            textColor=colors.HexColor("#0F172A")
        )
        
        subtitle_style = ParagraphStyle(
            'DocSubTitle',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=10,
            leading=14,
            textColor=colors.HexColor("#64748B")
        )

        elements = []

        # Organization Title & Header
        title_text = f"<b>{org_name or settings.ORGANIZATION_NAME}</b>"
        elements.append(Paragraph(title_text, title_style))
        elements.append(Paragraph(f"Official Face Recognition Attendance Log Report &bull; Generated on {datetime.now().strftime('%B %d, %Y at %I:%M %p')}", subtitle_style))
        elements.append(Spacer(1, 14))

        # KPI Summary Box Table
        total_recs = len(records)
        present_cnt = sum(1 for r in records if r.status.value == "PRESENT")
        late_cnt = sum(1 for r in records if r.status.value == "LATE")
        absent_cnt = sum(1 for r in records if r.status.value == "ABSENT")
        rate = ((present_cnt + late_cnt) / total_recs * 100) if total_recs > 0 else 0.0

        kpi_data = [
            ["Total Logs", "Present", "Late", "Absent", "Attendance Rate"],
            [str(total_recs), str(present_cnt), str(late_cnt), str(absent_cnt), f"{rate:.1f}%"]
        ]
        
        kpi_table = Table(kpi_data, colWidths=[140, 140, 140, 140, 160])
        kpi_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#F1F5F9")),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor("#0F172A")),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 1), (-1, 1), 12),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
        ]))
        elements.append(kpi_table)
        elements.append(Spacer(1, 16))

        # Records Table
        table_headers = [
            "ID", "Person ID", "Name", "Department", "Date", "Check-In", "Check-Out", "Status", "Confidence", "Mode"
        ]
        
        table_rows = [table_headers]
        for r in records:
            p = r.person
            dept = p.department.name if p and p.department else "N/A"
            table_rows.append([
                str(r.id),
                p.identifier if p else "-",
                p.full_name if p else "Unknown",
                dept,
                r.date.strftime("%Y-%m-%d") if r.date else "-",
                r.check_in_time.strftime("%I:%M %p") if r.check_in_time else "--",
                r.check_out_time.strftime("%I:%M %p") if r.check_out_time else "--",
                r.status.value,
                f"{r.recognition_confidence * 100:.0f}%",
                r.verification_mode.value[:4]
            ])

        col_widths = [35, 65, 130, 110, 75, 75, 75, 65, 55, 55]
        records_table = Table(table_rows, colWidths=col_widths, repeatRows=1)
        
        t_style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1E293B")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (2, 1), (3, -1), 'LEFT'), # Left align Name & Dept
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8.5),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
        ]

        # Alternating background colors and status styling
        for i, r in enumerate(records, 1):
            if i % 2 == 0:
                t_style.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor("#F8FAFC")))
            if r.status.value == "PRESENT":
                t_style.append(('TEXTCOLOR', (7, i), (7, i), colors.HexColor("#166534")))
                t_style.append(('FONTNAME', (7, i), (7, i), 'Helvetica-Bold'))
            elif r.status.value == "LATE":
                t_style.append(('TEXTCOLOR', (7, i), (7, i), colors.HexColor("#92400E")))
                t_style.append(('FONTNAME', (7, i), (7, i), 'Helvetica-Bold'))
            elif r.status.value == "ABSENT":
                t_style.append(('TEXTCOLOR', (7, i), (7, i), colors.HexColor("#991B1B")))
                t_style.append(('FONTNAME', (7, i), (7, i), 'Helvetica-Bold'))

        records_table.setStyle(TableStyle(t_style))
        elements.append(records_table)

        doc.build(elements)
        return buffer.getvalue()

report_service = ReportService()
